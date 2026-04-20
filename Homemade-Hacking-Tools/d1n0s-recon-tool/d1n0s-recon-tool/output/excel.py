import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(color="FFFFFF", bold=True)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def _style_header(ws):
    for cell in ws[1]:
        cell.fill  = HEADER_FILL
        cell.font  = HEADER_FONT
        cell.alignment = HEADER_ALIGN


def _autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)


def write(results: dict, path: str = "Attack_Surface.xlsx"):

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _write_companies(wb, results)
    _write_ranges(wb, results)
    _write_domains(wb, results)
    _write_subdomains(wb, results)

    if os.path.exists(path):
        ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
        base, ext = os.path.splitext(path)
        path = f"{base}_{ts}{ext}"

    wb.save(path)
    print(f"[+] Excel guardado en: {path}")


# Hojas

def _write_companies(wb, results):
    ws = wb.create_sheet("Empresas")
    headers = [
        "Nombre", "Fuente", "RIRs", "BGP HE",
        "WhoisXMLAPI - Reverse Netblock", "WhoisXMLAPI - Reverse Whois", "Cloud_enum"
    ]
    ws.append(headers)
    for company in results.get("companies", []):
        ws.append([
            company.get("name", ""),
            company.get("source", ""),
            company.get("rirs", ""),
            company.get("bgp_he", ""),
            company.get("whoisxml_netblock", ""),
            company.get("whoisxml_whois", ""),
            company.get("cloud_enum", ""),
        ])
    _style_header(ws)
    _autofit(ws)


def _write_ranges(wb, results):
    ws = wb.create_sheet("Rangos de red")
    headers = ["ASN", "Nombre", "Fuente", "País"]
    ws.append(headers)
    for r in results.get("asns", []):
        ws.append([
            r.get("asn", ""),
            r.get("name", ""),
            r.get("source", ""),
            r.get("country", ""),
        ])
    _style_header(ws)
    _autofit(ws)


def _write_domains(wb, results):
    ws = wb.create_sheet("Dominios")
    headers = [
        "Dominio", "Fuente", "Amass - Intel",
        "ViewDNS - Reverse NS", "ViewDNS - Reverse MX",
        "Amass - Enum", "Assetfinder", "Shodan - SSL",
        "Subscan", "Check SecurityTrails", "Check leaks"
    ]
    ws.append(headers)
    for d in results.get("domains", []):
        ws.append([
            d.get("domain", ""),
            d.get("source", ""),
            d.get("amass_intel", ""),
            d.get("reverse_ns", ""),
            d.get("reverse_mx", ""),
            d.get("amass_enum", ""),
            d.get("assetfinder", ""),
            d.get("shodan_ssl", ""),
            d.get("subscan", ""),
            d.get("securitytrails", ""),
            d.get("leaks", ""),
        ])
    _style_header(ws)
    _autofit(ws)
    
def _write_subdomains(wb, results):
    ws = wb.create_sheet("Subdominios")
    headers = ["FQDN", "Dominio Padre", "Dirección IP", "Resolución", "EyeWitness", "Nuclei"]
    ws.append(headers)
    for s in results.get("subdomains", []):
        ws.append([
            s.get("fqdn", ""),
            s.get("parent", ""),
            s.get("ip", ""),
            s.get("resolves", ""),
            s.get("eyewitness", ""),
            s.get("nuclei", ""),
        ])
    _style_header(ws)
    _autofit(ws)